import os
import pandas as pd
import snowflake.connector as snow
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import yaml
from datetime import datetime
from difflib import get_close_matches
import glob
import pdb
import numpy as np

def loadYaml():
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
    return config

def load_query_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

class DynamicSchemaComparator:
    def __init__(self, db2_file_path, snowflake_file_path):
        self.db2_file_path = db2_file_path
        self.snowflake_file_path = snowflake_file_path

    def load_data(self):
        if not os.path.exists(self.db2_file_path):
            raise FileNotFoundError(f"DB2 file not found: {self.db2_file_path}")
        if not os.path.exists(self.snowflake_file_path):
            raise FileNotFoundError(f"Snowflake file not found: {self.snowflake_file_path}")
        self.db2_df = pd.read_excel(self.db2_file_path)
        self.snowflake_df = pd.read_excel(self.snowflake_file_path)
        

    def map_columns(self):
        self.column_mapping = {}
        db2_columns = self.db2_df.columns.astype(str).str.strip().str.lower()
        snowflake_columns = self.snowflake_df.columns.astype(str).str.strip().str.lower()

        for col in db2_columns:
            match = get_close_matches(col, snowflake_columns, n=1, cutoff=0.6)
            if match:
                self.column_mapping[col] = match[0]
            else:
                print(f"Warning: No close match found for DB2 column '{col}'")

    def clean_data(self, df):
        df = df.map(lambda x: '-' if pd.isna(x) or (isinstance(x, str) and x.strip() == '') else x)
        df = df.map(lambda x: str(x).strip() if isinstance(x, str) else x)
        return df

    def compare_schemas(self):
        self.map_columns()  # Ensure that columns are mapped between DB2 and Snowflake.
        common_columns = self.column_mapping.keys()  # Get the common columns to compare.

        # Get the original column names in both dataframes
        db2_columns_original = self.db2_df.columns
        snowflake_columns_original = self.snowflake_df.columns

        # Convert column names to lowercase for comparison
        db2_columns = db2_columns_original.astype(str).str.strip().str.lower()
        snowflake_columns = snowflake_columns_original.astype(str).str.strip().str.lower()

        # Create a dictionary to map original column names to lowercase column names
        db2_column_mapping = {col_lower: col_original for col_lower, col_original in zip(db2_columns, db2_columns_original)}
        snowflake_column_mapping = {col_lower: col_original for col_lower, col_original in zip(snowflake_columns, snowflake_columns_original)}

        # Sort the dataframes by original column names
        self.db2_df = self.db2_df.sort_values(by=db2_columns_original.tolist())
        self.snowflake_df = self.snowflake_df.sort_values(by=snowflake_columns_original.tolist())

        comparison_results = {}

        max_length = max(len(self.db2_df), len(self.snowflake_df))

        for db2_col in common_columns:
            snowflake_col = self.column_mapping[db2_col]

            # Get the original column names from the lowercase column names
            db2_col_original = db2_column_mapping[db2_col.strip().lower()]
            snowflake_col_original = snowflake_column_mapping[snowflake_col.strip().lower()]

            db2_cleaned = self.clean_data(self.db2_df[[db2_col_original]])
            snowflake_cleaned = self.clean_data(self.snowflake_df[[snowflake_col_original]])

            # Align indexes before comparing
            db2_cleaned = db2_cleaned.reset_index(drop=True)
            snowflake_cleaned = snowflake_cleaned.reset_index(drop=True)

            # Pad the Series to the maximum length
            db2_padded = np.pad(db2_cleaned[db2_col_original].values, (0, max_length - len(db2_cleaned)))
            snowflake_padded = np.pad(snowflake_cleaned[snowflake_col_original].values, (0, max_length - len(snowflake_cleaned)))

            # Store DB2 and Snowflake data in the results dictionary
            comparison_results[f'{db2_col_original}_DB2'] = db2_padded
            comparison_results[f'{snowflake_col_original}_Snowflake'] = snowflake_padded

            # Compare the two columns after aligning indexes
            comparison_results[f'{db2_col_original}_Comparison'] = (db2_padded == snowflake_padded)

        # Convert the results into a DataFrame
        comparison_df = pd.DataFrame(comparison_results)

        # Fill missing values with NaN
        comparison_df = comparison_df.apply(lambda x: x.fillna(np.nan))

        return comparison_df
    def save_results(self, comparison_df, output_file_prefix='dynamic_schema_comparison_result'):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{output_file_prefix}_{timestamp}.xlsx"
        comparison_df.to_excel(output_file, index=False)

        timestamp= datetime.now.strftime('%Y%m%d_%H%M%S')
        # Load the workbook and the first sheet
        workbook = openpyxl.load_workbook(output_file)
        sheet = workbook.active

        # Define a yellow fill for highlighting
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and isinstance(header, str) and header.endswith('_Comparison'):
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is True:
                        cell.fill = yellow_fill

        workbook.save(output_file)
        print(f"Comparison results saved to {output_file}")

class profiling:
        
    def __init__(self, runID, runtime):
        self.runId = runID
        self.runTime = runtime
        self.config = loadYaml()
        self.loadConfig()
        self.fetch_db2_data_source()
        self.fetch_db2_data_target()
        self.find_latest_files()
        self.compare_schemas()

    def loadConfig(self):
        # DB2 Source Config
        self.db2_source_hostname = self.config["source_conn"]["hostname"]
        self.db2_source_uid = self.config["source_conn"]["uid"]
        self.db2_source_pwd = os.getenv(self.config["source_conn"]["pwd"])
        self.db2_source_database = self.config["source_conn"]["database"]
        self.db2_source_port = self.config["source_conn"]["port"]
        self.dll_variable = self.config["target_conn2"]["dll_path"]
        os.add_dll_directory(self.dll_variable)

        # DB2 Target Config
        self.db2_target_hostname = self.config["target_conn"]["hostname"]
        self.db2_target_uid = self.config["target_conn"]["uid"]
        self.db2_target_pwd = os.getenv(self.config["target_conn"]["pwd"])
        self.db2_target_database = self.config["target_conn"]["database"]
        self.db2_target_port = self.config["target_conn"]["port"]

    def fetch_db2_data_source(self):
        try:
            # DB2 Source Connection
            dsn_source = (
                "DRIVER={{IBM DB2 ODBC DRIVER}};"
                f"DATABASE={self.db2_source_database};"
                f"HOSTNAME={self.db2_source_hostname};"
                f"PORT={self.db2_source_port};"
                f"PROTOCOL=TCPIP;"
                f"UID={self.db2_source_uid};"
                f"PWD={self.db2_source_pwd};"
            )
            conn_source = ibm_db.connect(dsn_source, "", "")
            
            # Load query from file for source DB2
            query_text_source = load_query_from_file('db2_query_source.txt')

            stmt_source = ibm_db.exec_immediate(conn_source, query_text_source)
            result_source = []
            row = ibm_db.fetch_tuple(stmt_source)
            while row:
                result_source.append(row)
                row = ibm_db.fetch_tuple(stmt_source)

            df_db2_source = pd.DataFrame(result_source, columns=[
                'TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME', 'ORDINAL_POSITION', 'DB2_TYPE', 'DERIVED_LENGTH', 'SCALE',
                'COLUMN_DEFAULT', 'IS_NULLABLE'
            ])

            self.db2_source_file_path = self.get_incremented_filename('db2_source_data', 'xlsx')
            df_db2_source.to_excel(self.db2_source_file_path, index=False)

        except Exception as e:
            print(f"An error occurred while fetching source DB2 data: {e}")

        finally:
            if conn_source:
                ibm_db.close(conn_source)

    def fetch_db2_data_target(self):
        try:
            # DB2 Target Connection
            dsn_target = (
                "DRIVER={{IBM DB2 ODBC DRIVER}};"
                f"DATABASE={self.db2_target_database};"
                f"HOSTNAME={self.db2_target_hostname};"
                f"PORT={self.db2_target_port};"
                f"PROTOCOL=TCPIP;"
                f"UID={self.db2_target_uid};"
                f"PWD={self.db2_target_pwd};"
            )
            conn_target = ibm_db.connect(dsn_target, "", "")
            
            # Load query from file for target DB2
            query_text_target = load_query_from_file('db2_query_target.txt')

            stmt_target = ibm_db.exec_immediate(conn_target, query_text_target)
            result_target = []
            row = ibm_db.fetch_tuple(stmt_target)
            while row:
                result_target.append(row)
                row = ibm_db.fetch_tuple(stmt_target)

            df_db2_target = pd.DataFrame(result_target, columns=[
                'TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME', 'ORDINAL_POSITION', 'DB2_TYPE', 'DERIVED_LENGTH', 'SCALE',
                'COLUMN_DEFAULT', 'IS_NULLABLE'
            ])

            self.db2_target_file_path = self.get_incremented_filename('db2_target_data', 'xlsx')
            df_db2_target.to_excel(self.db2_target_file_path, index=False)

        except Exception as e:
            print(f"An error occurred while fetching target DB2 data: {e}")

        finally:
            if conn_target:
                ibm_db.close(conn_target)
            

    def get_incremented_filename(self, base_name, ext):
        i = 1
        while True:
            filename = f"{base_name}_{i}.{ext}"
            if not os.path.exists(filename):
                return filename
            i += 1

    def find_latest_files(self):
        # Fetch the latest DB2 file
        db2_files = glob.glob('db2_data_*.xlsx')
        if db2_files:
            self.db2_file_path = max(db2_files, key=os.path.getmtime)
        else:
            raise FileNotFoundError("No DB2 files found.")

        # Fetch the latest Snowflake file
        snowflake_files = glob.glob('snowflake_data_*.xlsx')
        if snowflake_files:
            self.snowflake_file_path = max(snowflake_files, key=os.path.getmtime)
        else:
            raise FileNotFoundError("No Snowflake files found.")

        print(f"Latest DB2 file: {self.db2_file_path}")
        print(f"Latest Snowflake file: {self.snowflake_file_path}")

    def compare_schemas(self):
        comparator = DynamicSchemaComparator(self.db2_file_path, self.snowflake_file_path)
        comparator.load_data()
        comparison_result = comparator.compare_schemas()
        comparator.save_results(comparison_result)

# Usage
profiling_instance = profiling(runID='test_run', runtime='test_time')
